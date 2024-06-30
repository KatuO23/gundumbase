import urllib.parse,time,requests,os
import openpyxl
from bs4 import BeautifulSoup

source_url = 'https://www.gundam-base.net/products/'
html = requests.get(source_url).text
url_list = []
target_li = 'li.next > a'
cell_x = 1#エクセルの行
cell_y = 1#エクセルの列

#対象のwebサイトはページネーションがあるためすべてのページのURLを取得する
def get_next_url(target_li,html):
    next_flag = True
    soup = BeautifulSoup(html,'html5lib')
    if source_url not in url_list:
        url_list.append(source_url)
    while next_flag:
        url = soup.select(target_li)
        if url == []:
            next_flag = False
        else:
            next_url = urllib.parse.urljoin(source_url,url[0].attrs['href'])
            url_list.append(next_url)
            #次のページのhtmlの解析
            html = requests.get(next_url).text
            soup = BeautifulSoup(html,'html5lib')
            time.sleep(1)
    return url_list

#商品情報を取得する
def get_item_info(target_li,html):
    html_list = []#htmlのテキストを入れる
    item_name_list = [] #商品名
    specWrap_info = []#商品情報
    item_info_list = [] #商品名と商品情報を入れる変数
    item_id_list = [] #devのidプロパティ情報を入れる変数
    for url in get_next_url(target_li,html):
        html = requests.get(url).text
        html_list.append(html)
    for html in html_list:
        soup = BeautifulSoup(html,'html5lib')
        for item in soup.select('p.name'):
            item.parent.parent.attrs['id']
            #すでに配列にある商品名は入れない
            #新商品のお知らせの部分の情報は入れない
            if item.text not in item_name_list and 'new' not in item.parent.parent.attrs['id']:
                item_name_list.append(item.text)
                item_id_list.append(item.parent.parent.attrs['id'])
        #商品情報を配列に入れる
        for item in item_id_list:
            specWrap_info.append(soup.select_one(f'#{item}>a>div>.specWrap').text.replace(' ','').replace('\n','').replace('\t',''))
            soup.select_one(f'#{item}>a>p.name').text
        #2ページ以降のページのidを検索するため配列をクリアする。
        item_id_list.clear()    
    #商品情報と商品名を紐づける    
    item_info_list = dict(zip(item_name_list,specWrap_info))
    #print(item_info_list)
    return item_info_list


def read_excel(cell_x,cell_y):
    out_dir = 'gundumbase'
    book = openpyxl.Workbook()
    sheet = book.active
    for key ,value in get_item_info(target_li,html).items():
        sheet.cell(row=cell_y,column=cell_x,value=key)
        sheet.cell(row=cell_y,column=cell_x+1,value=value)
        cell_y = cell_y+1
    #指定したフォルダーがない場合は作成する。
    if not os.path.exists(out_dir):
        os.mkdir(out_dir)
    savefile = os.path.join(out_dir,'gundumbase_info.xlsx')
    book.save(savefile)    

if __name__ == '__main__':
    read_excel(cell_x,cell_y)
    #get_item_info(target_li,html)